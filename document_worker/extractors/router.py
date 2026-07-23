from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import fitz
from docx import Document
from openpyxl import load_workbook


def extract_pdf(path: Path) -> str:
    with fitz.open(path) as document:
        return "\n".join(page.get_text("text") for page in document)


def extract_docx(path: Path) -> str:
    document = Document(path)
    return "\n".join(paragraph.text for paragraph in document.paragraphs)


def extract_xlsx(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"[시트] {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(value) for value in row if value is not None]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


def extract_hwpx(path: Path) -> str:
    fragments: list[str] = []
    with zipfile.ZipFile(path) as archive:
        section_names = sorted(
            name for name in archive.namelist() if name.startswith("Contents/section") and name.endswith(".xml")
        )
        for name in section_names:
            root = ElementTree.fromstring(archive.read(name))
            fragments.extend(element.text or "" for element in root.iter() if element.text)
    return "\n".join(fragments)


def extract_text(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        text = extract_pdf(path)
    elif suffix == ".docx":
        text = extract_docx(path)
    elif suffix in {".xlsx", ".xlsm"}:
        text = extract_xlsx(path)
    elif suffix == ".hwpx":
        text = extract_hwpx(path)
    elif suffix in {".txt", ".md", ".csv"}:
        text = path.read_text(encoding="utf-8", errors="replace")
    else:
        raise ValueError(f"지원하지 않는 문서 형식입니다: {suffix}")
    return re.sub(r"\n{3,}", "\n\n", text).strip()
